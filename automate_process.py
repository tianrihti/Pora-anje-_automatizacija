import pandas as pd
import logging
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
import win32com.client
import time
import pythoncom
import xlwings as xw
import psutil


# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelAutomation:
    def __init__(self):
        self.pregled_file = os.path.abspath('43.xls')
        self.porocanje_file = os.path.abspath("poročanje proizvodnje2025.xlsm")
        self.plan_file = os.path.abspath("plan brizganja 2025 mesečni.xlsx")

        # Validate files exist
        self._validate_files()

    def _validate_files(self):
        """Validate that all required files exist"""
        files = [self.pregled_file, self.porocanje_file, self.plan_file]
        for file in files:
            if not os.path.exists(file):
                raise FileNotFoundError(f"File not found: {file}")
        
    def step1_copy_pregled_data(self):
        """Step 1: Copy data from Pregled.xls"""
        logger.info("Step 1: Copying data from Pregled.xls")
        
        try:
            # Read Pregled.xls
            try:
                df = pd.read_excel(self.pregled_file, sheet_name="Sheet1", engine='xlrd')
            except ImportError:
                df = pd.read_excel(self.pregled_file, sheet_name="Sheet1", engine='openpyxl')
            
            logger.info(f"Successfully read Pregled.xls with {len(df)} rows")
            return df
            
        except Exception as e:
            logger.error(f"Error reading Pregled.xls: {e}")
            raise
    
    def step2_paste_to_porocanje(self, data_df):
        """Step 2: Paste data into poročanje proizvodnje2025.xlsm sheet 'prilepi gosoft' using openpyxl (safe for macros)"""
        logger.info("Step 2: Pasting data into poročanje proizvodnje2025.xlsm (safe method)")
        try:
            from openpyxl import load_workbook

            # Load the workbook with macros preserved
            wb = load_workbook(self.porocanje_file, keep_vba=True)
            if 'prilepi gosoft' not in wb.sheetnames:
                ws = wb.create_sheet('prilepi gosoft')
            else:
                ws = wb['prilepi gosoft']
            
            # Clear the sheet
            ws.delete_rows(1, ws.max_row)

            # Write headers
            for c_idx, col_name in enumerate(data_df.columns, 1):
                ws.cell(row=1, column=c_idx, value=col_name)

            # Write DataFrame to sheet
            for r_idx, row in enumerate(data_df.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(self.porocanje_file)
            wb.close()
            logger.info(f"Successfully pasted {len(data_df)} rows to 'prilepi gosoft' sheet (safe method)")
            return True
        except Exception as e:
            logger.error(f"Error pasting data to poročanje proizvodnje2025.xlsm (safe method): {e}")
            raise
    
    def get_target_date(self) -> datetime:
        """Get the target date based on the rules"""
        today = datetime.now()
        if today.weekday() == 0:  # Monday
            # Go back to Friday
            target_date = today - timedelta(days=3)
        else:
            # Go back to yesterday
            target_date = today - timedelta(days=1)
        return target_date

    def step3_find_date_in_plan(self):
        """Step 3: Find the correct date in plan sheet"""
        logger.info("Step 3: Finding date in plan sheet")
        
        target_date = self.get_target_date()
        logger.info(f"Looking for date: {target_date.strftime('%Y-%m-%d')}")
        
        try:
            wb = load_workbook(self.plan_file, data_only=True)
            ws = wb["plan"]
            
            # Check row 4 for dates
            date_found = False
            target_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=4, column=col).value
                if cell_value:
                    # Try to parse as date
                    if isinstance(cell_value, datetime):
                        cell_date = cell_value.date()
                    else:
                        try:
                            cell_date = pd.to_datetime(cell_value).date()
                        except:
                            continue
                    
                    if cell_date == target_date.date():
                        target_col = col
                        date_found = True
                        logger.info(f"Found target date at column {col}")

                        # Check if the cell below contains "Fiksno"
                        fiksno_cell = ws.cell(row=5, column=col).value
                        if fiksno_cell == "Fiksno":
                            logger.info("Found 'Fiksno' below the target date")
                        else:
                            raise ValueError("Plan is not fixed yet")
                        
                        break
        
            if not date_found:
                raise ValueError(f"Target date {target_date.strftime('%Y-%m-%d')} not found in row 4")
            
            return target_col
            
        except Exception as e:
            logger.error(f"Error finding date in plan: {e}")
            raise

    def step4_copy_plan_range(self, start_col):
        """Step 4: Copy range from plan sheet"""
        logger.info(f"Step 4: Copying range from column {start_col}")

        try:
            wb = load_workbook(self.plan_file, data_only=True)  # Use data_only=True to get values
            ws = wb["plan"]

            copied_data = []  # Will store list of rows, each row is a list of 3 cell values

            for row in range(6, 45):  # Rows 6 to 44 inclusive
                row_data = []
                for col_offset in range(3):  # Copy 3 columns: start_col, start_col+1, start_col+2
                    cell = ws.cell(row=row, column=start_col + col_offset)
                    row_data.append(cell.value)  # This will get the calculated value, not the formula
                copied_data.append(row_data)

            logger.info(f"Copied range from column {start_col} to {start_col+2}, rows 6 to 44")
            return copied_data

        except Exception as e:
            logger.error(f"Error copying plan range: {e}")
            raise
    
    def step5_paste_to_brizganje(self, copied_data):
        logger.info("Step 5: Pasting data into 'brizganje izračun' sheet")
        try:
            # Kill any existing Excel processes
            self.kill_excel_processes()

            # Load workbook
            wb = load_workbook(self.porocanje_file, keep_vba=True)
            ws = wb["brizganje izračun"]

            # Get the target date
            """Get the target date based on the rules"""
            today = datetime.now()
            if today.weekday() == 0 or today.weekday() == 1:  # Monday or Tuesday
                # Go back to Thursday or Friday
                target_date = today - timedelta(days=4)
            else:
                # Go back to yesterday
                target_date = today - timedelta(days=2)
            logger.info(f"Looking for date: {target_date.strftime('%Y-%m-%d')}")

            # Search row 4, starting from column D (index 4)
            target_col = None
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=4, column=col).value
                if cell_value:
                    if isinstance(cell_value, datetime):
                        cell_date = cell_value.date()
                    else:
                        try:
                            cell_date = pd.to_datetime(cell_value).date()
                        except:
                            continue

                    if cell_date == target_date.date():
                        target_col = col
                        logger.info(f"Found target date in column {col}")
                        break

            if target_col is None:
                raise ValueError(f"Target date {target_date.strftime('%Y-%m-%d')} not found in row 4")

            paste_col = target_col - 1  # One column to the left

            logger.info(f"Pasting values into column {paste_col}")

            # Clear the target range
            clear_range = ws[f"{ws.cell(row=4, column=paste_col).column_letter}4:{ws.cell(row=4+len(copied_data), column=paste_col+2).column_letter}{4+len(copied_data)}"]
            for row in clear_range:
                for cell in row:
                    cell.value = None

            # Paste copied_data into brizganje izracun sheet
            start_row = 4  # Assuming we start from row 4 (just like when copying)
            for i, row_data in enumerate(copied_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=start_row + i, column=paste_col + j, value=value)
                    
            # Save the workbook
            wb.save(self.porocanje_file)
            wb.close()
            logger.info("Successfully pasted values to 'brizganje izračun' sheet")
            return  # If successful, exit the function

        except Exception as e:
            logger.error(f"Error in Step 5: {e}")
            raise

    def step6_analyze_brizganje(self):
        try:
            logger.info("Step 6: Analyzing rows 7 to 46 in 'brizganje izračun'")

            # Load workbook with formula results
            wb = load_workbook(self.porocanje_file, data_only=True)
            ws = wb["brizganje izračun"]

            saved_texts = []
            
            logger.setLevel(logging.DEBUG)
            for row in range(7, 47): # Rows 7 to 46 inclusive
                cell_a = ws.cell(row=row, column=1).value  # Column A
                if not cell_a:
                    continue  # Skip if column A is empty

                cell_l = ws.cell(row=row, column=12).value  # Column L (12)
                if cell_l in (None, ""):
                    continue  # Skip if column L is empty or zero

                cell_m = ws.cell(row=row, column=13).value  # Column M (13)
                # Try converting to float if possible
                try:
                    if isinstance(cell_m, str):
                        # Remove euro sign and replace comma with dot, if needed
                        cleaned = cell_m.replace("€", "").replace(",", ".").strip()
                        value_m = float(cleaned)
                    else:
                        value_m = float(cell_m)
                except (TypeError, ValueError):
                    continue  # Skip if not a number

                if value_m > 50:
                    saved_texts.append(str(cell_a))  # Save value from column A

            logger.info(f"Saved texts: {saved_texts}")
            return saved_texts

        except Exception as e:
            logger.error(f"Error in Step 6: {e}")
            raise

    def recalc_excel(self):
        max_retries = 3
        retry_delay = 5  # seconds

        for attempt in range(max_retries):
            try:
                logger.info(f"Recalculating Excel (Attempt {attempt + 1})")
                self.kill_excel_processes()  # Ensure no Excel processes are running

                app = xw.App(visible=False)
                wb = app.books.open(self.porocanje_file)
                
                logger.info("Calculating...")
                wb.app.calculate()
                
                logger.info("Saving workbook...")
                wb.save()
                
                logger.info("Closing workbook...")
                wb.close()
                
                logger.info("Quitting Excel application...")
                app.quit()
                
                logger.info("Excel recalculation completed successfully")
                return  # If successful, exit the function

            except Exception as e:
                logger.error(f"Error in recalc_excel (Attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    logger.info(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                else:
                    logger.error("Max retries reached. Unable to recalculate Excel.")
                    raise
            finally:
                try:
                    if 'app' in locals():
                        app.quit()
                except:
                    pass
                self.kill_excel_processes()

        # If we've exhausted all retries, raise an exception
        raise Exception("Failed to recalculate Excel after multiple attempts")

    def kill_excel_processes(self):
        logger.info("Attempting to kill all Excel processes")
        for proc in psutil.process_iter(['name']):
            try:
                if proc.info['name'].lower() in ['excel.exe', 'xlview.exe']:
                    proc.kill()
                    logger.info(f"Killed Excel process: {proc.pid}")
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        time.sleep(2)  # Wait for 2 seconds after killing processes

    def step7_process_saved_texts(self, saved_texts):
        logger.info("Step 7: Processing saved texts")
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(self.porocanje_file)
            izbor_sheet = wb.Worksheets("izbor")
            list2_sheet = wb.Worksheets("List2")
            brizganje_izracun_sheet = wb.Worksheets("brizganje izračun")  
            
            # Delete all existing shapes (images) in the range A7:M44
            for shape in brizganje_izracun_sheet.Shapes:
                if (7 <= shape.TopLeftCell.Row <= 44 and 
                    1 <= shape.TopLeftCell.Column <= 13): 
                    shape.Delete()

            for text in saved_texts:
                logger.info(f"Processing text: {text}")

                # a. Filter data in Izbor sheet manually
                last_row = izbor_sheet.Cells(izbor_sheet.Rows.Count, "F").End(-4162).Row
                filtered_data = []

                for row in range(1, last_row + 1):
                    cell_value = izbor_sheet.Cells(row, 27).Value  # Column AA
                    if cell_value == text or row == 1:  # Include header row
                        row_data = [izbor_sheet.Cells(row, col).Value for col in range(6, 14)]  # Columns F to M
                        filtered_data.append(row_data)

                # b. Clear List2 target range
                last_row_list2 = list2_sheet.Cells(list2_sheet.Rows.Count, "T").End(-4162).Row
                list2_sheet.Range(f"T2:AA{last_row_list2}").ClearContents()

                # c. Paste filtered data into List2 sheet
                for i, row_data in enumerate(filtered_data):
                    for j, value in enumerate(row_data):
                        list2_sheet.Cells(i + 1, 20 + j).Value = value  # Start from column T (20th column)

                # d. Execute macro (gumb1)
                excel.Run("sortiraj")

                # Step 8: Copy processed data
                self.step8_copy_processed_data(list2_sheet)

                # Step 9: Paste as image in "brizganje izracun" sheet
                self.step9_paste_as_image(brizganje_izracun_sheet, text)

                excel.CutCopyMode = False  # Clear clipboard

            # Set the height of rows 7 to 44 to 16.5 if they don't contain an image
            for row in range(7, 45):
                has_image = False
                for shape in brizganje_izracun_sheet.Shapes:
                    if shape.TopLeftCell.Row == row:
                        has_image = True
                        break
                if not has_image:
                    brizganje_izracun_sheet.Rows(row).RowHeight = 16.5
            logger.info("Adjusted heights of rows without images to 16.5")

            wb.Save()
            wb.Close()
            excel.Quit()
            logger.info("Step 7,8 and 9 completed successfully")

        except Exception as e:
            logger.error(f"Error in Step 7: {e}")
            raise
        finally:
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
            self.kill_excel_processes()
    
    def enable_macros(self, wb):
        try:
            # Check if there's a security alert
            if wb.ReadOnly:
                # Try to enable content
                app = wb.Application
                app.DisplayAlerts = False
                app.EnableEvents = False
                app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
                app.Run("Auto_Open")  # This usually enables content
                app.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception as e:
            logger.error(f"Failed to enable macros: {e}")
    
    def step8_copy_processed_data(self, list2_sheet):
        """Step 8: Copy processed data from List2 sheet"""
        logger.info("Step 8: Copying processed data from List2")

        # Find the last row with data in column C
        last_row = 8  # Default to row 8
        for row in range(9, 28):  # Check rows 9 and 10
            if list2_sheet.Cells(row, 3).Value:  # Column C
                last_row = row

        # Construct the range to copy
        range_to_copy = list2_sheet.Range(f"B1:L{last_row}")
        range_to_copy.CopyPicture(Appearance=1, Format=2)  # Copy as picture

        logger.info(f"Successfully copied range B1:L{last_row} from List2 as picture")

    def step9_paste_as_image(self, brizganje_izracun_sheet, text):
        """Step 9: Paste as image in 'brizganje izračun' sheet"""
        logger.info(f"Step 9: Pasting as image for text '{text}'")

        last_row_brizganje = brizganje_izracun_sheet.Cells(brizganje_izracun_sheet.Rows.Count, "A").End(-4162).Row
        target_row = None
        for row in range(1, last_row_brizganje + 1):
            if brizganje_izracun_sheet.Cells(row, 1).Value == text:
                target_row = row + 1  # Go one row below
                break

        if target_row:
            target_cell = brizganje_izracun_sheet.Cells(target_row, 1)
            brizganje_izracun_sheet.Paste(target_cell, Link=False)
    
            # Get the last pasted shape (which should be our image)
            last_shape = brizganje_izracun_sheet.Shapes(brizganje_izracun_sheet.Shapes.Count)
    
            # Adjust row height to fit the image
            image_height = last_shape.Height
            brizganje_izracun_sheet.Rows(target_row).RowHeight = image_height
    
            logger.info(f"Successfully pasted image for text '{text}' at row {target_row} and adjusted row height")
            # Add this line to check if the shape is actually there
            logger.info(f"Shape count after pasting: {brizganje_izracun_sheet.Shapes.Count}")
        else:
            logger.warning(f"Could not find row for text '{text}' in 'brizganje izračun' sheet")


if __name__ == "__main__":
    # Create automation instance
    automation = ExcelAutomation()
    try:
        automation.kill_excel_processes()
        # Run step 1
        pregled_data = automation.step1_copy_pregled_data()

        # Run step 2
        automation.step2_paste_to_porocanje(pregled_data)

        # Run step 3
        target_col = automation.step3_find_date_in_plan()

        # Run step 4
        plan_range_data = automation.step4_copy_plan_range(target_col)

        # Run step 5
        automation.step5_paste_to_brizganje(plan_range_data)

        automation.recalc_excel()
        # Run step 6
        saved_texts = automation.step6_analyze_brizganje()

        # Run step 7
        automation.step7_process_saved_texts(saved_texts)
    except Exception as e:
        logger.error((f"An error occurred: {e}"))
    finally:
        automation.kill_excel_processes()
        logger.info("Script execution finished")